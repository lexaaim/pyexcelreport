#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
import itertools

from openpyxl.utils import get_column_letter
from .xlsutils_apply import *
from .xlscolor import *

from recordclass import recordclass

class XLSTableField:
    """Структура для хранения информации одного столбца данных таблицы
    """
    def __init__(self, fieldname, format='string', col_count=1, editable=False, hidden=False, default_value=None):
        self.fname            = fieldname
        self.ccount           = col_count
        self.format           = format
        self.editable         = editable
        self.hidden           = hidden
        self.default_value    = default_value

FieldStruct = recordclass('FieldStruct', 'findex xls_start xls_end format hidden '
                                         'last_value last_value_row changed '
                                         'hide_condition hide_flag '
                                         'merging subtitle subtitle_rowcount subtotal '
                                         'color_fn pattern_fn')

class XLSTable:
    """Класс, инкапсулирующий информацию и методы отображения данных таблицы
    """
    def __init__(self, colinfo, data, row_height=30):
        if len(data) > 0:
            assert len(colinfo) == len(data[0]), "количество полей в структуре таблицы не совпадает с фактическим количеством"

        self._fields = dict()
        findex = 0
        cindex = 0
        for ci in colinfo:
            self._fields[ci.fname] = FieldStruct(
                        findex=findex,
                        xls_start=cindex, xls_end=cindex + ci.ccount - 1,
                        format=ci.format, hidden=ci.hidden,
                        last_value=None, last_value_row=None, changed=False,
                        hide_condition=None, hide_flag=None,
                        merging=False, subtitle=None, subtitle_rowcount=0, subtotal=None,
                        color_fn=None, pattern_fn=None)
            findex += 1
            if not ci.hidden:
                cindex += ci.ccount

        self._data = data
        self._hierarchy = []
        self._row_height = row_height
        self._col_count = sum([ci.ccount for ci in colinfo if not ci.hidden])
        self._row_count = len(data)
        self._calculate_fn = None

    def add_hide_column_condition(self, fieldname, cond_func):
        self._fields[fieldname].hide_condition = cond_func
        self._fields[fieldname].hide_flag = True

    def add_coloring(self, fieldname, color_fn):
        self._fields[fieldname].color_fn = color_fn

    def add_pattern(self, fieldname, pattern_fn):
        self._fields[fieldname].pattern_fn = pattern_fn

    def set_calculating(self, calc_func):
        self._calculate_fn = calc_func

    def hierarchy_append(self, fieldname, merging=False, subtitle=None, subtotal=None, subtitle_rowcount=0):
        self._hierarchy.append(fieldname)
        self._fields[fieldname].merging = merging
        self._fields[fieldname].subtitle = subtitle
        self._fields[fieldname].subtotal = subtotal
        self._fields[fieldname].subtitle_rowcount = subtitle_rowcount

    def group_by_data(self, colinfo, hierarchy, sums):
        # TODO: asserts
        hier_indexes = [self._fields[fname].findex for fname in hierarchy]
        key_func = lambda d: tuple([(d[i] is None, d[i]) for i in hier_indexes])

        table_data = copy(self._data)
        table_data.sort(key=key_func)

        table_total_data=[]
        for key, rows in itertools.groupby(table_data, key_func):
            stored_rows = list(rows)
            data_row = []
            for ci in colinfo:
                if ci.fname in hierarchy:
                    data_row.append(key[hierarchy.index(ci.fname)][1])
                elif ci.fname in sums:
                    cindex = self._fields[ci.fname].findex
                    # TODO: refactor to zip
                    data_row.append(sum(r[cindex] for r in stored_rows))
            table_total_data.append(tuple(data_row))
        return table_total_data

    def get_column_xls_index_pair(self, fieldname):
        return (self._fields[fieldname].xls_start, self._fields[fieldname].xls_end)

    def apply(self, ws, first_row, first_col):
        """Отображает непосредственно в XLS данные таблицы
        """
        def _before_line_processing(row):
            """ставим флаг changed если значение поля в структуре hierarchy поменяло свое значение
            """
            was_changed = False
            # ordering sensitive
            fieldlist = [fn for fn in self._hierarchy]
            for fieldname in fieldlist:
                f = self._fields[fieldname]
                if (row is None) or (row[f.findex] != f.last_value):
                    was_changed = True
                if was_changed:
                   self._fields[fieldname].changed = True

        def _after_line_processing(row, cur_row):
            """сохраняем инфо о последних значениях для всех полей в структуре _hierarchy
            """
            fieldlist = [fn for fn in self._hierarchy if self._fields[fn].changed]
            for fieldname in fieldlist:
                f = self._fields[fieldname]
                self._fields[fieldname].last_value = row[f.findex] if row else None
                self._fields[fieldname].last_value_row = cur_row
                self._fields[fieldname].changed = False

        def _merge_previous_row(cur_row):
            """объединяем ячейки с одинаковыми значениями, и те, что шириной более 1 столбца
            """
            #TODO: refactor
            fields = [f for _, f in self._fields.items()
                        if (not f.merging) and (f.xls_end - f.xls_start != 0)]
            for f in fields:
                apply_range(ws, cur_row - 1, first_col + f.xls_start,
                                cur_row - 1, first_col + f.xls_end, set_merge)
                apply_range(ws, cur_row - 1, first_col + f.xls_start,
                                cur_row - 1, first_col + f.xls_end, set_borders)

            fields = [self._fields[fn] for fn in self._hierarchy if self._fields[fn].merging]
            for f in fields:
                if (f.changed) and (f.last_value_row is not None) and (f.last_value_row != cur_row - 1):
                    apply_range(ws, f.last_value_row, first_col + f.xls_start,
                                    cur_row - 1,      first_col + f.xls_end, set_merge)
                    apply_range(ws, f.last_value_row, first_col + f.xls_start,
                                    cur_row - 1,      first_col + f.xls_end, set_borders)
                elif (f.changed) and (f.xls_end - f.xls_start != 0):
                    apply_range(ws, cur_row - 1, first_col + f.xls_start,
                                    cur_row - 1, first_col + f.xls_end, set_merge)
                    apply_range(ws, cur_row - 1, first_col + f.xls_start,
                                    cur_row - 1, first_col + f.xls_end, set_borders)

        def _make_subtotals(cur_row):
            """делаем подитоги
            """
            stlines = 0
            # ordering sensitive
            fields = [self._fields[fn] for fn in reversed(self._hierarchy) if self._fields[fn].subtotal and self._fields[fn].changed]
            for fch in fields:
                if (fch.last_value_row is not None) and (fch.last_value_row != cur_row - 1):
                    _row, fchcol = cur_row + stlines, first_col + fch.xls_start

                    ws.row_dimensions[_row].height = 18

                    # если подитоги по невидимому столбцу, печатаем заголовок в первом столбце
                    # TODO: refactor
                    _label_col = fchcol if not fch.hidden else first_col
                    if not fch.hidden and fch.xls_start != fch.xls_end:
                        apply_range(ws, _row, first_col + fch.xls_start,
                                        _row, first_col + fch.xls_end, set_merge)
                    ws.cell(row=_row, column=_label_col).value = "Σ '{0:s}'".format(str(fch.last_value))
                    apply_cell(ws, _row, _label_col, set_alignment)
                    apply_cell(ws, _row, _label_col, set_font, bold=True)

                    for st in fch.subtotal:
                        f = self._fields[st]
                        fcol1, fcol2 = first_col + f.xls_start, first_col + f.xls_end
                        if (fcol1 - fcol2 > 1):
                            apply_range(ws, _row, fcol1, _row, fcol2, set_merge)

                        formulae = "=SUBTOTAL(9,{0:s}{1:d}:{0:s}{2:d})".format(
                                get_column_letter(fcol1), fch.last_value_row, cur_row - 1)
                        ws.cell(row=_row, column=fcol1).value = formulae
                        apply_range(ws, _row, fcol1, _row, fcol2, set_merge)

                        apply_range(ws, _row, fcol1, _row, fcol2, set_alignment, horizontal='right')
                        apply_range(ws, _row, fcol1, _row, fcol2, set_format, format=f.format)
                        apply_range(ws, _row, fcol1, _row, fcol2, set_borders)
                        apply_range(ws, _row, fcol1, _row, fcol2, set_fill, color=Color.LT_GRAY.value)

                    for i in range(fch.last_value_row - fch.subtitle_rowcount, cur_row + stlines):
                        ws.row_dimensions[i].outlineLevel += 1
                    stlines += 1

            return cur_row + stlines

        def _make_headers(cur_row, data_row):
            # ordering sensitive
            fields = [self._fields[fn] for fn in self._hierarchy if self._fields[fn].subtitle and self._fields[fn].changed]

            #TODO: change base container type???
            if fields:
                data_dict_row = dict()
                for fn, f in self._fields.items():
                    data_dict_row[fn] = data_row[f.findex]

                for fch in fields:
                    cur_row = fch.subtitle(ws, data_dict_row, cur_row, first_col)

            return cur_row

        def _calculate_fields(data_row):
            #TODO: change base container type???
            if self._calculate_fn:
                data_dict_row = dict()
                for fn, f in self._fields.items():
                    data_dict_row[fn] = data_row[f.findex]

                if self._calculate_fn:
                    self._calculate_fn(data_dict_row)

                    for fn, f in data_dict_row.items():
                        data_row[self._fields[fn].findex] = data_dict_row[fn]

        def _coloring(cur_row, data_row):
            #TODO: change base container type???
            data_dict_row = dict()
            for fn, f in self._fields.items():
                data_dict_row[fn] = data_row[f.findex]

            fields = [f for f in self._fields.values() if f.color_fn]
            for f in fields:
                col = f.color_fn(data_dict_row)
                apply_range(ws, cur_row, first_col + f.xls_start,
                                cur_row, first_col + f.xls_end,
                                set_fill, color=col.value)

            fields = [f for f in self._fields.values() if f.pattern_fn]
            for f in fields:
                (pattern, colbg, colfg) = f.pattern_fn(data_dict_row)
                apply_range(ws, cur_row, first_col + f.xls_start,
                                cur_row, first_col + f.xls_end,
                                set_pattern_fill, bg_color=colbg.value, fg_color=colfg.value, pattern_type=pattern)

        cur_row = first_row
        data_row_number = 0
        for data_row in self._data:
            sys.stdout.write("\rИдёт форматирование таблицы {0:0=2d}%".format(data_row_number * 100 // self._row_count ))
            sys.stdout.flush()

            _calculate_fields(data_row)
            _before_line_processing(data_row)
            if cur_row > first_row:
                _merge_previous_row(cur_row)
            cur_row = _make_subtotals(cur_row)
            cur_row = _make_headers(cur_row, data_row)
            _coloring(cur_row, data_row)

            ws.row_dimensions[cur_row].height = self._row_height

            for fieldname, f in self._fields.items():
                if f.hidden: continue

                #  if (f.xls_end != f.xls_start):
                #      apply_range(ws, cur_row, first_col + f.xls_start,
                #                      cur_row, first_col + f.xls_end, set_merge)

                # если печатаю числа, не выводить нулевые значения
                if (f.format != 'empty'):
                    if (f.format not in ['int', '1digit', 'currency', '3digit']) or (data_row[f.findex] != 0):
                        ws.cell(row=cur_row, column=first_col + f.xls_start).value = data_row[f.findex]

                # обновляем флаг hide_flag чтобы скрыть в конце неиспользуемые колонки
                if (f.hide_condition is not None) and (not f.hide_condition(data_row[f.findex])):
                    self._fields[fieldname].hide_flag = False

            _after_line_processing(data_row, cur_row)

            # apply format and alignment
            for f in self._fields.values():
                if f.hidden: continue

                xlr = get_xlrange(cur_row, first_col + f.xls_start, cur_row, first_col + f.xls_end)
                if f.format in ['int', 'currency', '1digit', '3digit']:
                    apply_xlrange(ws, xlr, set_alignment, horizontal='right')
                    apply_xlrange(ws, xlr, set_format, format=f.format)
                elif f.format in ['date']:
                    apply_xlrange(ws, xlr, set_alignment)
                    apply_xlrange(ws, xlr, set_format, format=f.format)
                else:
                    apply_xlrange(ws, xlr, set_alignment)

            # apply borders, outline, font
            cr = get_xlrange(cur_row, first_col, cur_row, first_col + self._col_count - 1)
            apply_xlrange(ws, cr, set_borders)
            apply_xlrange(ws, cr, set_font)

            cur_row += 1
            data_row_number += 1

        sys.stdout.write("\n")

        _before_line_processing(None)
        _merge_previous_row(cur_row)
        cur_row = _make_subtotals(cur_row)

        # скрываем все колонки, для которых не выполнились условия
        fields = [[f.xls_start, f.xls_end] for f in self._fields.values() if f.hide_flag and not f.hidden]
        for fstart, fend in fields:
            for i in range(fstart, fend + 1):
                ws.column_dimensions[get_column_letter(first_col + i)].hidden = True

        # apply borders, outline, font
        cr = get_xlrange(first_row, first_col, cur_row - 1, first_col + self._col_count - 1)
        apply_xlrange(ws, cr, set_outline, border_style='medium')

        return cur_row
